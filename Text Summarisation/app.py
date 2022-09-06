from flask import Flask,render_template,request
from rouge import Rouge
import bert_score
from bert_score import score
import matplotlib.pyplot as plt
from bert_score import BERTScorer
import pandas as pd
from pandas import ExcelWriter
from werkzeug.utils import secure_filename
from summarizer import Summarizer, TransformerSummarizer
from nltk.translate.bleu_score import sentence_bleu
import numpy
import tensorflow as tf 
##Print many staement at same time using the below command
##from IPython.core.interactiveshell import InteractiveShell
##InteractiveShell.ast_node_interactivity = "all"
import fitz
import glob
import pdfplumber
import re
import os
#import xlsxwriter
import json
from openpyxl import load_workbook

df = pd.DataFrame()

r = 0
col = 0

app = Flask(__name__)
UPLOAD_FOLDER = './uploads'
app.config['data'] = UPLOAD_FOLDER
F1_Bert=''
F1_Gpt2=''

def listToString(s):
   
    # initialize an empty string
    str1 = ""
   
    # traverse in the string 
    for ele in s:
        str1 += ele 
   
    # return string 
    return str1
    
    
def clean_text(text):
    
    text = re.sub(r"[^A-Za-z0-9.]+", ' ', text)
    return text

"""def text_extract():
    print("+++++++++++++++++++++++++++++++++++++++++++",'HI')
    pdf_files = glob.glob("data/" + "/*.pdf")
    org_text = ""
    all_abstracts = ""
    for filename in pdf_files:
        print("+++++++++++++++++++++++++++++++++++++++++++",filename)
        doc = fitz.open(filename)
        #print ("number of pages: %i" % doc.pageCount)
        #print (doc.is_repaired)
        #print (doc.page_count)
        ##for page in doc:
        ##    page1text = page.getText()
        ##    page2text = page.getText()

        page1 = doc.load_page(0)
        page1text = page1.get_text("text")
        
        page2 = doc.load_page(1)
        page2text = page2.get_text("text")

        try:
            try:
                Abstract = page1text.split("Introduction",1)[1]
                partitioned_string = page2text.split('Data')[0]
            except:
                Abstract = page1text.split("Introduction",1)[1]
                partitioned_string = Abstract.split('Data')[0]
        except Exception as IndexError:
            try:
                Abstract = page2text.split("Introduction",1)[1]
                partitioned_string = Abstract.split('Data')[0]
            except:
                Abstract = page2text.split("Introduction",1)[1]
                partitioned_string = Abstract.split('Data')[0]
        partitioned_string = re.sub('\S+@\S+',"",partitioned_string)
        partitioned_string = re.sub(r'[0-9]',"",partitioned_string)
        #partitioned_string = " ".join(re.findall(r"[a-zA-Z0-9]+",partitioned_string ))
        all_abstracts = partitioned_string
        org_text=all_abstracts
        dic[filename] = org_text
    
   
    df = pd.DataFrame(dic.items(), columns=['Filename', 'Abstract'])
    writer = ExcelWriter('Data.xlsx')
    df.to_excel(writer,sheet_name = 'Filename')
    writer.save()
    
      
    return (df)"""



@app.route('/')
def index():
    dic = {}
    df = pd.read_excel('Data.xlsx')
    abstracts = df['Abstract'].tolist()
    l = len(abstracts)
    for i in range(l):
        #abstracts[i] = re.sub(r"(@\[A-Za-z0-9]+)|([^0-9A-Za-z \t])|(\w+:\/\/\S+)|^rt|http.+?", " ", abstracts[i])
        abstracts[i] = abstracts[i]
        
    filename = df['Filename'].tolist()
    for i,j in zip(filename,abstracts):
        dic[i] = [j]
    #print(dic.keys())
    return render_template('datamart.html', original = dic)

   
@app.route('/bert', methods = ['GET', 'POST'])
def index1():
    df = pd.read_excel('Data.xlsx')
    abstracts = df['Abstract'].tolist()
    filename = df['Filename'].tolist()
    my_dict = {}
    my_bert = {}
    my_dictionary = {}
    bert_model = Summarizer()
    j=0
    for i in abstracts:
        
        bert_summary = ''.join(bert_model(i))
        file = open('org.txt', 'w', encoding='utf-8')
        file.write(i)
        file.close()
        blue_score = {}
        file = open('hyps.txt', 'w', encoding='utf-8')
        file.write(bert_summary)
        file.close()
        r = Rouge()
        score_bert=r.get_scores(bert_summary, i)
        score_b = json.dumps(score_bert)
        with open("hyps.txt",encoding='utf-8') as f:
            cands = [line.strip() for line in f]

        with open("org.txt", encoding='utf-8') as f:
            refs = [[line.strip() for line in f]]
        blue_score_bert=sentence_bleu(refs, cands)
        
        #TODO
        if len(cands)==1:
            P, R, F1 = score(cands, refs, lang='en', rescale_with_baseline=True)
            F1_bert=F1
            
        if len(cands)>1:
            scorer = BERTScorer(lang="en", rescale_with_baseline=True)
            P, R, F1 = scorer.score(cands, refs)
            F1_bert=F1
        bert_=tf.constant(F1_bert).numpy()
    
        blue_score['Blue Score'] = F1_bert;
        print('BERT OUTPUT');
        print(F1_bert);
        print(bert_);
        df_bert = pd.DataFrame(score_bert[0]['rouge-1'], index=['Scores']).apply(lambda x: round(x,3))
        df_all_rough1= pd.concat([df_bert]) 
        df_all_rough1.columns.name = 'ROUGE-1'
        df_all_rough1.index.name = 'BERT'
        print(df_all_rough1[['f','p','r']])
    
        df_bert = pd.DataFrame(score_bert[0]['rouge-2'], index=['Scores']).apply(lambda x: round(x,3))
        df_all_rough2= pd.concat([df_bert]) 
        df_all_rough2.columns.name = 'ROUGE-2'
        df_all_rough2.index.name = 'BERT'
    
        df_bert = pd.DataFrame(score_bert[0]['rouge-l'], index=['Scores']).apply(lambda x: round(x,3))
        df_all_rough3= pd.concat([df_bert]) 
        df_all_rough3.columns.name = 'ROUGE-L'
        df_all_rough3.index.name = 'BERT'
        my_dictionary[bert_summary] = [filename[j]]
        my_dict[score_b] = [blue_score]
        my_bert[score_b] = bert_
        j = j + 1
    print(my_dict.keys())
    df1 = pd.DataFrame(my_dict.items(), columns=['Abstracts', 'BERT/Rouge Scores'])
    ExcelWorkbook = load_workbook('Data.xlsx') 
 
    # Generating the writer engine
    writer = pd.ExcelWriter('Data.xlsx', engine = 'openpyxl')
 
    # Assigning the workbook to the writer engine
    writer.book = ExcelWorkbook
    df1.to_excel(writer, sheet_name='Bert Scores')
    writer.save()
    print('FINAL');
    print(my_bert);
    return render_template('log.html',blue = my_dict , summ = my_dictionary,bert=my_bert)


@app.route('/gpt2', methods = ['GET', 'POST'])
def index2():
    df = pd.read_excel('Data.xlsx')
    abstracts = df['Abstract'].tolist()
    filename = df['Filename'].tolist()
    my_dict = {}
    my_dictionary = {}
    GPT2_model = TransformerSummarizer(transformer_type="GPT2",transformer_model_key="gpt2-medium")
    j = 0 
    for i in abstracts:
        gpt2_summ = ''.join(GPT2_model(i))
        r_gpt2 = Rouge()
        blue_score = {}
        score_gpt2=r_gpt2.get_scores(gpt2_summ, i)
        score_b = json.dumps(score_gpt2)
        file = open('hyps_gpt2.txt', 'w', encoding='utf-8')
        file.write(gpt2_summ)
        file.close()
        with open("hyps_gpt2.txt", encoding='utf-8') as f:
            cands_gpt2 = [line.strip() for line in f]
        with open("org.txt", encoding='utf-8') as f:
            refs = [[line.strip() for line in f]]
   
        blue_score_gpt2=sentence_bleu(refs, cands_gpt2)
        blue_score['Blue Score'] = blue_score_gpt2 
    #blue_score_gpt2=sentence_bleu(refs, cands_gpt2)
        df_gpt2 = pd.DataFrame(score_gpt2[0]['rouge-1'], index=['Scores']).apply(lambda x: round(x,3))
        df_all_rough1= pd.concat([df_gpt2]) 
        df_all_rough1.columns.name = 'ROUGE-1'
        df_all_rough1.index.name = 'GPT 2'
        print(df_all_rough1[['f','p','r']])

        df_gpt2 = pd.DataFrame(score_gpt2[0]['rouge-2'], index=['Scores']).apply(lambda x: round(x,3))
        df_all_rough2= pd.concat([df_gpt2]) 
        df_all_rough2.columns.name = 'ROUGE-2'
        df_all_rough2.index.name = 'GPT 2'
    
        df_gpt2 = pd.DataFrame(score_gpt2[0]['rouge-l'], index=['Scores']).apply(lambda x: round(x,3))
        df_all_rough3= pd.concat([df_gpt2]) 
        df_all_rough3.columns.name = 'ROUGE-L'
        df_all_rough3.index.name = 'GPT 2'
        my_dictionary[gpt2_summ] = [filename[j]]
        my_dict[score_b] = [blue_score]
        j = j + 1
    
    
    df1 = pd.DataFrame(my_dict.items(), columns=['Summary', 'GPT2/Rouge Scores'])
    ExcelWorkbook = load_workbook('Data.xlsx')
 
    # Generating the writer engine
    writer = pd.ExcelWriter('Data.xlsx', engine = 'openpyxl')
 
    # Assigning the workbook to the writer engine
    writer.book = ExcelWorkbook
    df1.to_excel(writer, sheet_name='GPT2 Scores')
    writer.save()
   
   
    return render_template('log.html', blue = my_dict, summ = my_dictionary)

#host="0.0.0.0", port=8002,

@app.route('/bert_all', methods = ['GET', 'POST'])
def index3():
    df = pd.read_excel('Data.xlsx')
    abstracts = df['Abstract'].tolist()
    filename = df['Filename'].tolist()
    my_dict = {}
    my_dictionary = {}
    bert_model = Summarizer()
    
    str = ""
    for i in abstracts:
        
        str = str + i 
    
    bert_summary = ''.join(bert_model(str))
    file = open('org.txt', 'w', encoding='utf-8')
    file.write(str)
    file.close()
    blue_score = {}
    file = open('hyps.txt', 'w', encoding='utf-8')
    file.write(bert_summary)
    file.close()
    r = Rouge()
    score_bert=r.get_scores(bert_summary, str)
    score_b = json.dumps(score_bert)
    with open("hyps.txt",encoding='utf-8') as f:
        cands = [line.strip() for line in f]
    
    
        
    with open("org.txt", encoding='utf-8') as f:
        refs = [[line.strip() for line in f]]
    blue_score_bert=sentence_bleu(refs, cands)
    blue_score['Blue Score'] = blue_score_bert
    df_bert = pd.DataFrame(score_bert[0]['rouge-1'], index=['Scores']).apply(lambda x: round(x,3))
    df_all_rough1= pd.concat([df_bert]) 
    df_all_rough1.columns.name = 'ROUGE-1'
    df_all_rough1.index.name = 'BERT'
    print(df_all_rough1[['f','p','r']])
    
    df_bert = pd.DataFrame(score_bert[0]['rouge-2'], index=['Scores']).apply(lambda x: round(x,3))
    df_all_rough2= pd.concat([df_bert]) 
    df_all_rough2.columns.name = 'ROUGE-2'
    df_all_rough2.index.name = 'BERT'
    
    df_bert = pd.DataFrame(score_bert[0]['rouge-l'], index=['Scores']).apply(lambda x: round(x,3))
    df_all_rough3= pd.concat([df_bert]) 
    df_all_rough3.columns.name = 'ROUGE-L'
    df_all_rough3.index.name = 'BERT'
    my_dictionary['Summary'] = [bert_summary]
    my_dict[score_b] = [blue_score]
    
    print(my_dict.keys())
    df1 = pd.DataFrame(my_dict.items(), columns=['Abstracts', 'BERT/Rouge Scores'])
    ExcelWorkbook = load_workbook('Data.xlsx') 
 
    # Generating the writer engine
    writer = pd.ExcelWriter('Data.xlsx', engine = 'openpyxl')
 
    # Assigning the workbook to the writer engine
    writer.book = ExcelWorkbook
    df1.to_excel(writer, sheet_name='Bert Scores')
    writer.save()
    
    return render_template('log.html',blue = my_dict , summ = my_dictionary)

@app.route('/gpt2_all', methods = ['GET', 'POST'])
def index4():
    df = pd.read_excel('Data.xlsx')
    abstracts = df['Abstract'].tolist()
    filename = df['Filename'].tolist()
    my_dict = {}
    my_dictionary = {}
    GPT2_model = TransformerSummarizer(transformer_type="GPT2",transformer_model_key="gpt2-medium")
    str = "" 
    for i in abstracts:
        str = str + i

    gpt2_summ = ''.join(GPT2_model(str))
    r_gpt2 = Rouge()
    blue_score = {}
    score_gpt2=r_gpt2.get_scores(gpt2_summ, str)
    score_b = json.dumps(score_gpt2)
    file = open('hyps_gpt2.txt', 'w', encoding='utf-8')
    file.write(gpt2_summ)
    file.close()
    with open("hyps_gpt2.txt", encoding='utf-8') as f:
        cands_gpt2 = [line.strip() for line in f]
    with open("org.txt", encoding='utf-8') as f:
        refs = [[line.strip() for line in f]]
   
    blue_score_gpt2=sentence_bleu(refs, cands_gpt2)
    blue_score['Blue Score'] = blue_score_gpt2 
    #blue_score_gpt2=sentence_bleu(refs, cands_gpt2)
    df_gpt2 = pd.DataFrame(score_gpt2[0]['rouge-1'], index=['Scores']).apply(lambda x: round(x,3))
    df_all_rough1= pd.concat([df_gpt2]) 
    df_all_rough1.columns.name = 'ROUGE-1'
    df_all_rough1.index.name = 'GPT 2'
    print(df_all_rough1[['f','p','r']])

    df_gpt2 = pd.DataFrame(score_gpt2[0]['rouge-2'], index=['Scores']).apply(lambda x: round(x,3))
    df_all_rough2= pd.concat([df_gpt2]) 
    df_all_rough2.columns.name = 'ROUGE-2'
    df_all_rough2.index.name = 'GPT 2'
    
    df_gpt2 = pd.DataFrame(score_gpt2[0]['rouge-l'], index=['Scores']).apply(lambda x: round(x,3))
    df_all_rough3= pd.concat([df_gpt2]) 
    df_all_rough3.columns.name = 'ROUGE-L'
    df_all_rough3.index.name = 'GPT 2'
    my_dictionary['Summary'] = [gpt2_summ]
    my_dict[score_b] = [blue_score]
       
    
    
    df1 = pd.DataFrame(my_dict.items(), columns=['Summary', 'GPT2/Rouge Scores'])
    ExcelWorkbook = load_workbook('Data.xlsx')
 
    # Generating the writer engine
    writer = pd.ExcelWriter('Data.xlsx', engine = 'openpyxl')
 
    # Assigning the workbook to the writer engine
    writer.book = ExcelWorkbook
    df1.to_excel(writer, sheet_name='GPT2 Scores')
    writer.save()
   
   
    return render_template('log.html', blue = my_dict, summ = my_dictionary)


if __name__ == '__main__':
    app.run(debug = True,host="0.0.0.0", port=8004)


#,host="0.0.0.0", port=8004
